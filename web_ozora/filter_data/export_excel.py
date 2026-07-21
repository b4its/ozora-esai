import os
import sys
from datetime import datetime

import django
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'web_ozora.settings')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
django.setup()

from django.contrib.auth.models import User
from django.utils import timezone
from iot_service.models import ExperimentRoom, SensorData

USERNAME = 'tester'
TARGET_DATE = datetime(2026, 7, 20)

def get_querysets(user, target_date):
    start = timezone.make_aware(datetime(
        target_date.year, target_date.month, target_date.day, 0, 0, 0))
    end = timezone.make_aware(datetime(
        target_date.year, target_date.month, target_date.day, 23, 59, 59, 999999))
    experiments = ExperimentRoom.objects.filter(
        user=user, created__gte=start, created__lte=end)
    sensor_data = SensorData.objects.filter(
        user=user, created__gte=start, created__lte=end)
    return experiments, sensor_data

def write_header(ws, headers):
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

def export_experiments(ws, experiments):
    headers = ['ID', 'Nama Eksperimen', 'Suhu Target (°C)', 'Flow Speed (L/s)',
               'pH Target', 'Dibuat']
    write_header(ws, headers)
    for row_idx, exp in enumerate(experiments, 2):
        ws.cell(row=row_idx, column=1, value=exp.id)
        ws.cell(row=row_idx, column=2, value=exp.name)
        ws.cell(row=row_idx, column=3, value=exp.suhu)
        ws.cell(row=row_idx, column=4, value=exp.flow_speed)
        ws.cell(row=row_idx, column=5, value=exp.ph_target)
        ws.cell(row=row_idx, column=6, value=exp.created.strftime('%Y-%m-%d %H:%M:%S'))
    auto_width(ws)

def export_sensor_data(ws, sensor_data):
    headers = ['ID', 'Experiment', 'Device', 'Raw Light', 'Red', 'Green', 'Blue',
               'Temp', 'Lux', 'Dibuat']
    write_header(ws, headers)
    for row_idx, sd in enumerate(sensor_data, 2):
        ws.cell(row=row_idx, column=1, value=sd.id)
        ws.cell(row=row_idx, column=2, value=str(sd.experiment) if sd.experiment else '')
        ws.cell(row=row_idx, column=3, value=str(sd.device) if sd.device else '')
        ws.cell(row=row_idx, column=4, value=float(sd.raw_light) if sd.raw_light else None)
        ws.cell(row=row_idx, column=5, value=float(sd.red))
        ws.cell(row=row_idx, column=6, value=float(sd.green))
        ws.cell(row=row_idx, column=7, value=float(sd.blue))
        ws.cell(row=row_idx, column=8, value=float(sd.temp))
        ws.cell(row=row_idx, column=9, value=float(sd.lux))
        ws.cell(row=row_idx, column=10, value=sd.created.strftime('%Y-%m-%d %H:%M:%S'))
    auto_width(ws)

def main():
    try:
        user = User.objects.get(username=USERNAME)
    except User.DoesNotExist:
        print(f"User '{USERNAME}' tidak ditemukan.")
        sys.exit(1)

    experiments, sensor_data = get_querysets(user, TARGET_DATE)

    wb = openpyxl.Workbook()
    ws_exp = wb.active
    ws_exp.title = 'ExperimentRoom'
    export_experiments(ws_exp, experiments)

    ws_sensor = wb.create_sheet('SensorData')
    export_sensor_data(ws_sensor, sensor_data)

    filename = f'export_{USERNAME}_{TARGET_DATE.strftime("%Y%m%d")}.xlsx'
    wb.save(filename)
    print(f'Ekspor berhasil: {filename}')
    print(f'  ExperimentRoom: {experiments.count()} baris')
    print(f'  SensorData: {sensor_data.count()} baris')

if __name__ == '__main__':
    main()
